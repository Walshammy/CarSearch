from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import pandas as pd
import time
import os
from datetime import datetime
import logging
import hashlib
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle

class StreamlinedMasterScraper:
    def __init__(self):
        # Setup logging
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        self.logger = logging.getLogger(__name__)
        
        # Setup Chrome options
        self.chrome_options = Options()
        self.chrome_options.add_argument('--headless')
        self.chrome_options.add_argument('--no-sandbox')
        self.chrome_options.add_argument('--disable-dev-shm-usage')
        self.chrome_options.add_argument('--window-size=1920,1080')
        
        # URLs to scrape
        self.urls = {
            'Toyota 86': 'https://www.trademe.co.nz/a/motors/cars/toyota/86',
            'Subaru BRZ': 'https://www.trademe.co.nz/a/motors/cars/subaru/brz'
        }
        
        # Output directory (main folder - one level up from scripts)
        self.output_dir = r"C:\Users\james\Downloads\CarSearch"
        os.makedirs(self.output_dir, exist_ok=True)
        
        # OneDrive backup directory
        self.onedrive_dir = r"C:\Users\james\OneDrive - Silverdale Medical Limited\CarSearch"
        os.makedirs(self.onedrive_dir, exist_ok=True)
        
        # 86/BRZ dataset file (saved directly in main folder, no subfolder)
        self.master_file = os.path.join(self.output_dir, "86_BRZ_dataset.xlsx")
        
        # OneDrive backup file
        self.onedrive_file = os.path.join(self.onedrive_dir, "86_BRZ_dataset.xlsx")

    def generate_unique_id(self, title, location, year):
        """Generate a unique ID based on listing characteristics"""
        unique_string = f"{title}_{location}_{year}_{datetime.now().strftime('%Y%m%d')}"
        return hashlib.md5(unique_string.encode()).hexdigest()[:12].upper()

    def extract_listing_data(self, listing_element, car_model):
        """Extract comprehensive data from a single listing element"""
        try:
            data = {}
            
            # Get the full text content
            full_text = listing_element.text.strip()
            
            if not full_text or len(full_text) < 20:
                return None
            
            # Split into lines
            lines = [line.strip() for line in full_text.split('\n') if line.strip()]
            
            # Basic info
            data['title'] = lines[0] if lines else 'N/A'
            
            # Extract year
            year = 'N/A'
            if data['title'] != 'N/A':
                words = data['title'].split()
                for word in words:
                    if word.isdigit() and len(word) == 4 and 1900 < int(word) < 2030:
                        year = word
                        break
            data['year'] = year
            
            # Extract brand
            data['brand'] = car_model.split()[0]  # Toyota or Subaru
            
            # Extract mileage
            mileage_text = 'N/A'
            for line in lines:
                if 'km' in line.lower() and any(char.isdigit() for char in line):
                    mileage_text = line
                    break
            data['kms'] = mileage_text
            
            # Extract location
            location_text = 'N/A'
            for line in lines:
                if any(word in line.lower() for word in ['city', 'auckland', 'wellington', 'christchurch', 'hamilton', 'tauranga', 'dunedin']):
                    location_text = line
                    break
            data['location'] = location_text
            
            # Generate unique ID
            data['ID'] = self.generate_unique_id(data['title'], data['location'], data['year'])
            
            # Extract price information
            price_text = 'N/A'
            for line in lines:
                if '$' in line and any(char.isdigit() for char in line):
                    price_text = line
                    break
            data['price'] = price_text
            
            # Determine if it's an auction
            is_auction = any(word in full_text.lower() for word in ['auction', 'bid', 'reserve', 'ending'])
            data['is_auction'] = is_auction
            data['price_type'] = 'Auction' if is_auction else 'Buy Now'
            
            # Determine seller type
            is_dealer = any(word in full_text.lower() for word in ['dealer', 'motors', 'cars', 'auto', 'ltd', 'limited'])
            data['seller_type'] = 'Dealer' if is_dealer else 'Private'
            data['is_dealer'] = is_dealer
            
            # Additional fields
            data['car_model'] = car_model
            data['listing_url'] = 'N/A'
            data['listing_id'] = 'N/A'
            data['transmission'] = 'N/A'
            data['fuel_type'] = 'N/A'
            data['body_style'] = 'N/A'
            data['scrape_date'] = datetime.now().strftime('%Y-%m-%d')
            data['scrape_time'] = datetime.now().strftime('%H:%M:%S')
            data['last_seen'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            data['is_active'] = True
            
            return data
            
        except Exception as e:
            self.logger.error(f"Error extracting listing data: {e}")
            return None

    def scrape_car_listings(self, car_model, url):
        """Scrape listings for a specific car model"""
        self.logger.info(f"Starting scrape for {car_model}")
        
        driver = webdriver.Chrome(options=self.chrome_options)
        all_listings = []
        
        try:
            # Navigate to the page
            driver.get(url)
            self.logger.info(f"Navigated to: {url}")
            
            # Wait for page to load
            time.sleep(5)
            
            # Look for listings
            listings = driver.find_elements(By.CSS_SELECTOR, '.tm-motors-tier-one-search-card__listing-details-container')
            
            if not listings:
                self.logger.warning(f"No listings found for {car_model}")
                return all_listings
            
            self.logger.info(f"Found {len(listings)} listings for {car_model}")
            
            # Extract data from each listing
            for i, listing in enumerate(listings[:20]):  # Limit to first 20 listings
                try:
                    listing_data = self.extract_listing_data(listing, car_model)
                    if listing_data:
                        all_listings.append(listing_data)
                        self.logger.info(f"Extracted listing {i+1}: {listing_data.get('title', 'N/A')[:50]}...")
                except Exception as e:
                    self.logger.error(f"Error processing listing {i+1}: {e}")
                    continue
            
            self.logger.info(f"Successfully extracted {len(all_listings)} listings for {car_model}")
            
        except Exception as e:
            self.logger.error(f"Error scraping {car_model}: {e}")
        
        finally:
            driver.quit()
        
        return all_listings

    def load_existing_dataset(self):
        """Load existing master dataset if it exists"""
        if os.path.exists(self.master_file):
            try:
                df = pd.read_excel(self.master_file)
                self.logger.info(f"Loaded existing dataset with {len(df)} records")
                return df
            except Exception as e:
                self.logger.error(f"Error loading existing dataset: {e}")
                return pd.DataFrame()
        else:
            self.logger.info("No existing dataset found, creating new one")
            return pd.DataFrame()

    def update_dataset(self, new_data):
        """Update the master dataset with new data"""
        # Load existing data
        existing_df = self.load_existing_dataset()
        
        if existing_df.empty:
            # Create new dataset
            updated_df = pd.DataFrame(new_data)
        else:
            # Update existing dataset
            new_df = pd.DataFrame(new_data)
            
            # Mark existing listings as potentially inactive
            existing_df['is_active'] = False
            
            # Add new data
            updated_df = pd.concat([new_df, existing_df], ignore_index=True)
            
            # Remove duplicates based on ID, keeping the most recent
            updated_df = updated_df.drop_duplicates(subset=['ID'], keep='first')
        
        return updated_df

    def clean_and_format_data(self, df):
        """Clean and format data for proper Excel number formatting"""
        try:
            # Create a copy to avoid modifying original
            cleaned_df = df.copy()
            
            # Clean year column - convert to numbers if possible
            cleaned_df['year'] = cleaned_df['year'].apply(lambda x: self.clean_number(x, 'year'))
            
            # Clean mileage column - extract numbers from text like "50,000 km"
            cleaned_df['kms'] = cleaned_df['kms'].apply(lambda x: self.clean_mileage(x))
            
            # Clean price column - extract numbers from text like "$25,000"
            cleaned_df['price'] = cleaned_df['price'].apply(lambda x: self.clean_price(x))
            
            self.logger.info("Data cleaned and formatted for Excel")
            return cleaned_df
            
        except Exception as e:
            self.logger.error(f"Error cleaning data: {e}")
            return df

    def clean_number(self, value, column_type):
        """Clean a number value, return number if valid, otherwise return original"""
        if pd.isna(value) or value == 'N/A' or value == '':
            return value
        
        try:
            # Remove any non-digit characters except decimal point
            cleaned = str(value).replace(',', '').replace(' ', '')
            # Extract just the number part
            number_match = re.search(r'(\d+(?:\.\d+)?)', cleaned)
            if number_match:
                number = float(number_match.group(1))
                if column_type == 'year':
                    # For years, return as integer if it's a reasonable year
                    if 1900 <= number <= 2030:
                        return int(number)
                return number
        except:
            pass
        
        return value

    def clean_mileage(self, value):
        """Clean mileage value - extract number from text like '50,000 km'"""
        if pd.isna(value) or value == 'N/A' or value == '':
            return value
        
        try:
            # Look for number followed by 'km'
            match = re.search(r'([\d,]+)\s*km', str(value), re.IGNORECASE)
            if match:
                # Remove commas and convert to number
                number_str = match.group(1).replace(',', '')
                return int(number_str)
        except:
            pass
        
        return value

    def clean_price(self, value):
        """Clean price value - extract number from text like '$25,000'"""
        if pd.isna(value) or value == 'N/A' or value == '':
            return '-'
        
        try:
            # Look for $ followed by number
            match = re.search(r'\$([\d,]+)', str(value))
            if match:
                # Remove commas and convert to number
                number_str = match.group(1).replace(',', '')
                return int(number_str)
        except:
            pass
        
        return '-'

    def apply_conditional_formatting(self, filepath):
        """Apply beautiful conditional formatting to the Excel file"""
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Define beautiful colors
            header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")  # Dark slate gray
            header_font = Font(bold=True, color="FFFFFF")  # White bold text
            
            active_id_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")  # Light green for ID column
            inactive_id_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Light gray for inactive ID column
            inactive_row_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")  # Very light gray for inactive rows
            
            # Light brand-specific colors
            toyota_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")  # Very light cream for Toyota
            subaru_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # Very light blue for Subaru
            
            # Style the header row (row 1)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set default alignment for all data cells
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    column_letter = cell.column_letter
                    
                    # Set alignment based on column type
                    if column_letter in ['A', 'C', 'D', 'E', 'H', 'J', 'M']:  # ID, Year, Kms, Price, Is auction, Is dealer, Is active
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif column_letter in ['B', 'F', 'G', 'I', 'L', 'N', 'O', 'P', 'R', 'S', 'T', 'U']:  # Text columns
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif column_letter in ['K', 'Q']:  # Title and URL columns (long text)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Find the is_active and brand columns
            header_row = 1
            is_active_col = None
            brand_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=header_row, column=col).value == 'is_active':
                    is_active_col = col
                elif ws.cell(row=header_row, column=col).value == 'brand':
                    brand_col = col
            
            # Apply formatting based on is_active and brand columns
            for row in range(2, ws.max_row + 1):
                # Get brand for this row
                brand_value = None
                if brand_col:
                    brand_value = ws.cell(row=row, column=brand_col).value
                
                # Determine brand-specific fill
                brand_fill = None
                if brand_value == 'Toyota':
                    brand_fill = toyota_fill
                elif brand_value == 'Subaru':
                    brand_fill = subaru_fill
                
                # Apply formatting based on is_active column
                if is_active_col:
                    is_active_cell = ws.cell(row=row, column=is_active_col)
                    
                    if is_active_cell.value == True:
                        # Active listing - light green ID column + brand color for brand column
                        ws.cell(row=row, column=1).fill = active_id_fill  # ID column (column 1)
                        if brand_col and brand_fill:
                            ws.cell(row=row, column=brand_col).fill = brand_fill
                    else:
                        # Inactive listing - light gray ID column and slightly lighter row
                        ws.cell(row=row, column=1).fill = inactive_id_fill  # ID column
                        # Make the entire row slightly lighter
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = inactive_row_fill
                else:
                    # No is_active column, just apply brand colors
                    if brand_col and brand_fill:
                        ws.cell(row=row, column=brand_col).fill = brand_fill
            
            # Apply currency formatting to price column
            price_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == 'price':
                    price_col = col
                    break
            
            # Apply number formatting to appropriate columns
            for col in range(1, ws.max_column + 1):
                column_name = ws.cell(row=1, column=col).value
                
                if column_name == 'price':
                    # Currency formatting for price column
                    for row in range(2, ws.max_row + 1):
                        price_cell = ws.cell(row=row, column=col)
                        if isinstance(price_cell.value, (int, float)) and price_cell.value != '-':
                            price_cell.number_format = '$#,##0'  # Currency format with $ symbol
                
                elif column_name == 'year':
                    # Number formatting for year column
                    for row in range(2, ws.max_row + 1):
                        year_cell = ws.cell(row=row, column=col)
                        if isinstance(year_cell.value, (int, float)):
                            year_cell.number_format = '0'  # Integer format
                
                elif column_name == 'kms':
                    # Number formatting for mileage column
                    for row in range(2, ws.max_row + 1):
                        kms_cell = ws.cell(row=row, column=col)
                        if isinstance(kms_cell.value, (int, float)):
                            kms_cell.number_format = '#,##0'  # Number format with commas
            
            # Auto-adjust column widths with better formatting
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                # Calculate max length for this column
                for cell in column:
                    try:
                        if cell.value is not None:
                            # For headers, use the header text length
                            if cell.row == 1:
                                max_length = max(max_length, len(str(cell.value)))
                            else:
                                # For data cells, consider the actual content length
                                cell_length = len(str(cell.value))
                                max_length = max(max_length, cell_length)
                    except:
                        pass
                
                # Set minimum widths for different column types
                min_widths = {
                    'A': 8,   # ID column
                    'B': 10,  # Brand column
                    'C': 6,   # Year column
                    'D': 12,  # Kms column
                    'E': 12,  # Price column
                    'F': 20,  # Location column
                    'G': 15,  # Price type column
                    'H': 10,  # Is auction column
                    'I': 15,  # Seller type column
                    'J': 10,  # Is dealer column
                    'K': 30,  # Title column
                    'L': 15,  # Car model column
                    'M': 10,  # Is active column
                    'N': 15,  # Last seen column
                    'O': 15,  # Scrape date column
                    'P': 15,  # Scrape time column
                    'Q': 50,  # Listing URL column
                    'R': 15,  # Listing ID column
                    'S': 12,  # Transmission column
                    'T': 12,  # Fuel type column
                    'U': 15   # Body style column
                }
                
                # Get the minimum width for this column
                min_width = min_widths.get(column_letter, 10)
                
                # Calculate final width (max of calculated length + 2, minimum width, but cap at 60)
                final_width = max(max_length + 2, min_width)
                final_width = min(final_width, 60)  # Cap at 60 characters
                
                ws.column_dimensions[column_letter].width = final_width
                
                # Enable text wrapping for long content columns
                if column_letter in ['K', 'Q']:  # Title and URL columns
                    for cell in column:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            wb.save(filepath)
            self.logger.info("Applied beautiful conditional formatting to Excel file")
            
        except Exception as e:
            self.logger.error(f"Error applying conditional formatting: {e}")

    def save_master_dataset(self, df):
        """Save the master dataset with proper formatting"""
        if df.empty:
            self.logger.warning("No data to save")
            return
        
        # Reorder columns as requested: ID, brand, year, kms, price, location
        column_order = [
            'ID', 'brand', 'year', 'kms', 'price', 'location', 'price_type', 
            'is_auction', 'seller_type', 'is_dealer', 'title', 'car_model', 
            'is_active', 'last_seen', 'scrape_date', 'scrape_time', 
            'listing_url', 'listing_id', 'transmission', 'fuel_type', 'body_style'
        ]
        
        # Only keep columns that exist in the data
        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]
        
        # Sort by most recent first
        df = df.sort_values(['scrape_date', 'last_seen'], ascending=[False, False])
        
        # Clean and format data for proper Excel number formatting
        df = self.clean_and_format_data(df)
        
        # Save to Excel (no timestamped copies, just update the master file)
        filepath = self.master_file
        
        try:
            # Save main file
            df.to_excel(filepath, index=False, engine='openpyxl')
            self.logger.info(f"86/BRZ dataset saved to: {filepath}")
            
            # Apply conditional formatting to main file
            self.apply_conditional_formatting(filepath)
            
            # Save backup copy to OneDrive
            df.to_excel(self.onedrive_file, index=False, engine='openpyxl')
            self.logger.info(f"86/BRZ dataset backup saved to: {self.onedrive_file}")
            
            # Apply conditional formatting to OneDrive backup
            self.apply_conditional_formatting(self.onedrive_file)
            
            # Print summary
            print(f"\n=== 86/BRZ Dataset Summary ===")
            print(f"Total listings: {len(df)}")
            print(f"Active listings: {len(df[df['is_active'] == True])}")
            print(f"Inactive listings: {len(df[df['is_active'] == False])}")
            for model in df['car_model'].unique():
                count = len(df[df['car_model'] == model])
                print(f"{model}: {count} listings")
            print(f"Main dataset: {self.master_file}")
            print(f"OneDrive backup: {self.onedrive_file}")
            
        except Exception as e:
            self.logger.error(f"Error saving master dataset: {e}")

    def scrape_all_cars(self):
        """Scrape listings for all car models"""
        all_data = []
        
        for car_model, url in self.urls.items():
            listings = self.scrape_car_listings(car_model, url)
            all_data.extend(listings)
            time.sleep(2)  # Delay between requests
        
        return all_data

    def run(self):
        """Main execution method"""
        self.logger.info("Starting 86/BRZ Dataset Scraper")
        start_time = datetime.now()
        
        try:
            # Scrape all car data
            new_data = self.scrape_all_cars()
            
            # Update master dataset
            updated_df = self.update_dataset(new_data)
            
            # Save master dataset
            self.save_master_dataset(updated_df)
            
            end_time = datetime.now()
            duration = end_time - start_time
            self.logger.info(f"Master dataset update completed in {duration}")
            
        except Exception as e:
            self.logger.error(f"Error during master dataset update: {e}")

def main():
    """Main function to run the master scraper"""
    scraper = StreamlinedMasterScraper()
    scraper.run()

if __name__ == "__main__":
    main()
