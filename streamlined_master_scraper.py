from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import pandas as pd
import time
import os
from datetime import datetime, timedelta
import logging
import hashlib
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle

class StreamlinedMasterScraper:
    def __init__(self):
        # Setup logging
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        self.logger = logging.getLogger(__name__)
        
        # Setup Chrome options
        self.chrome_options = Options()
        self.chrome_options.add_argument('--headless')
        self.chrome_options.add_argument('--no-sandbox')
        self.chrome_options.add_argument('--disable-dev-shm-usage')
        self.chrome_options.add_argument('--window-size=1920,1080')
        
        # URLs to scrape
        self.urls = {
            'Toyota 86': 'https://www.trademe.co.nz/a/motors/cars/toyota/86',
            'Subaru BRZ': 'https://www.trademe.co.nz/a/motors/cars/subaru/brz'
        }
        
        # Output directory (main CarSearch folder)
        self.output_dir = r"C:\Users\james\Downloads\CarSearch"
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Daily backups directory
        self.daily_backups_dir = os.path.join(self.output_dir, "daily_backups")
        os.makedirs(self.daily_backups_dir, exist_ok=True)
        
        # OneDrive backup directory
        self.onedrive_dir = r"C:\Users\james\OneDrive - Silverdale Medical Limited\CarSearch"
        os.makedirs(self.onedrive_dir, exist_ok=True)
        
        # 86/BRZ dataset file (saved directly in main folder, no subfolder)
        self.master_file = os.path.join(self.output_dir, "86_BRZ_dataset.xlsx")
        
        # OneDrive backup file
        self.onedrive_file = os.path.join(self.onedrive_dir, "86_BRZ_dataset.xlsx")

    def generate_unique_id(self, title, location, year):
        """Generate a unique ID based on listing characteristics (without date to maintain consistency)"""
        unique_string = f"{title}_{location}_{year}"
        return hashlib.md5(unique_string.encode()).hexdigest()[:12].upper()

    def categorize_listing(self, year, price):
        """Categorize listing based on buying guide criteria"""
        try:
            # Convert year to int if possible
            year_int = None
            if year and year != 'N/A':
                try:
                    year_int = int(year)
                except (ValueError, TypeError):
                    pass
            
            # Convert price to int if possible
            price_int = None
            if price and price != 'N/A' and price != '':
                try:
                    if isinstance(price, str):
                        # Remove $ and commas
                        price_str = str(price).replace('$', '').replace(',', '')
                        price_int = int(price_str)
                    else:
                        price_int = int(price)
                except (ValueError, TypeError):
                    pass
            
            # Categorize based on year and price
            if year_int:
                if year_int in [2012, 2013]:
                    return 'avoid'  # Avoid 2012-2013 due to valve spring recall
                elif year_int in [2015, 2016]:
                    if price_int and 18000 <= price_int <= 23000:
                        return 'steal'  # Perfect: 2015-2016 in budget
                    else:
                        return 'optimal'  # 2015-2016 but outside budget
                elif year_int == 2014:
                    if price_int and 18000 <= price_int <= 23000:
                        return 'budget'  # 2014 in budget (backup option)
                    else:
                        return 'optimal'  # 2014 but outside budget
                else:
                    return 'other'  # Other years
            else:
                return 'other'  # No year data
                
        except Exception as e:
            self.logger.error(f"Error categorizing listing: {e}")
            return 'other'

    def generate_search_terms(self, title, location, year, brand, car_model):
        """Generate search terms and URLs to help find the original listing"""
        try:
            # Clean up the title for better search results
            clean_title = title.replace('N/A', '').strip()
            if not clean_title:
                clean_title = f"{brand} {car_model}"
            
            # Create various search term combinations
            search_terms = []
            
            # Basic search terms
            if year != 'N/A' and year:
                search_terms.append(f"{brand} {car_model} {year}")
                search_terms.append(f"{year} {brand} {car_model}")
            
            # Include location in search
            if location != 'N/A' and location:
                search_terms.append(f"{brand} {car_model} {location}")
                if year != 'N/A' and year:
                    search_terms.append(f"{brand} {car_model} {year} {location}")
            
            # Use original title if it's descriptive
            if clean_title and len(clean_title) > 10:
                search_terms.append(clean_title)
                if location != 'N/A' and location:
                    search_terms.append(f"{clean_title} {location}")
            
            # Remove duplicates and limit to 5 best search terms
            unique_terms = list(dict.fromkeys(search_terms))[:5]
            
            # Create TradeMe search URLs
            trademe_urls = []
            for term in unique_terms[:3]:  # Top 3 terms for TradeMe
                encoded_term = term.replace(' ', '+')
                trademe_url = f"https://www.trademe.co.nz/a/motors/cars/search?search_string={encoded_term}"
                trademe_urls.append(trademe_url)
            
            # Create Google search URLs
            google_urls = []
            for term in unique_terms[:3]:  # Top 3 terms for Google
                encoded_term = term.replace(' ', '+')
                google_url = f"https://www.google.com/search?q={encoded_term}+site:trademe.co.nz"
                google_urls.append(google_url)
            
            # Create Google Images search URLs
            google_images_urls = []
            for term in unique_terms[:2]:  # Top 2 terms for Google Images
                encoded_term = term.replace(' ', '+')
                google_images_url = f"https://www.google.com/search?q={encoded_term}+site:trademe.co.nz&tbm=isch"
                google_images_urls.append(google_images_url)
            
            return {
                'search_terms': ' | '.join(unique_terms),
                'trademe_search_urls': ' | '.join(trademe_urls),
                'google_search_urls': ' | '.join(google_urls),
                'google_images_urls': ' | '.join(google_images_urls),
                'primary_search_term': unique_terms[0] if unique_terms else f"{brand} {car_model}",
                'primary_trademe_url': trademe_urls[0] if trademe_urls else '',
                'primary_google_url': google_urls[0] if google_urls else ''
            }
            
        except Exception as e:
            self.logger.error(f"Error generating search terms: {e}")
            return {
                'search_terms': f"{brand} {car_model}",
                'trademe_search_urls': '',
                'google_search_urls': '',
                'google_images_urls': '',
                'primary_search_term': f"{brand} {car_model}",
                'primary_trademe_url': '',
                'primary_google_url': ''
            }

    def extract_listing_data(self, listing_element, car_model):
        """Extract comprehensive data from a single listing element with intelligent parsing"""
        try:
            data = {}
            
            # Get the full text content
            full_text = listing_element.text.strip()
            
            if not full_text or len(full_text) < 20:
                return None
            
            # Split into lines
            lines = [line.strip() for line in full_text.split('\n') if line.strip()]
            
            # Initialize notes for any unclear data
            notes = []
            
            # Extract title (first line, but clean it up)
            title = lines[0] if lines else 'N/A'
            data['title'] = title
            
            # Extract year from title using regex
            year = 'N/A'
            if title != 'N/A':
                year_match = re.search(r'\b(19|20)\d{2}\b', title)
                if year_match:
                    year = year_match.group()
            data['year'] = year
            
            # Extract brand
            data['brand'] = car_model.split()[0]  # Toyota or Subaru
            
            # Intelligent mileage extraction
            mileage_text = 'N/A'
            mileage_found = False
            
            for line in lines:
                # Look for patterns like "50,000 km", "50000km", "Low kms", etc.
                mileage_patterns = [
                    r'(\d{1,3}(?:,\d{3})*)\s*km',  # 50,000 km
                    r'(\d+)\s*km',                   # 50000 km
                    r'(\d{1,3}(?:,\d{3})*)km',      # 50,000km
                    r'(\d+)km'                      # 50000km
                ]
                
                for pattern in mileage_patterns:
                    match = re.search(pattern, line, re.IGNORECASE)
                    if match:
                        mileage_text = match.group(1).replace(',', '')
                        mileage_found = True
                        break
                
                if mileage_found:
                    break
            
            # If no mileage found, check for "low km" indicators
            if not mileage_found:
                for line in lines:
                    if re.search(r'\b(low|super low|very low)\s*km\b', line, re.IGNORECASE):
                        mileage_text = 'Low km'
                        break
            
            data['kms'] = mileage_text
            
            # Intelligent price extraction
            price_text = 'N/A'
            for line in lines:
                # Look for $ followed by numbers
                price_match = re.search(r'\$(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', line)
                if price_match:
                    price_value = price_match.group(1).replace(',', '')
                    try:
                        price_num = float(price_value)
                        # Only keep prices >= $1000, otherwise leave blank
                        if price_num >= 1000:
                            price_text = price_match.group(0)
                        else:
                            price_text = ''  # Blank for prices < $1000
                    except ValueError:
                        price_text = price_match.group(0)  # Keep original if can't parse
                    break
            
            data['price'] = price_text
            
            # Intelligent location extraction
            location_text = 'N/A'
            nz_cities = ['auckland', 'wellington', 'christchurch', 'hamilton', 'tauranga', 'dunedin', 
                        'palmerston north', 'napier', 'hastings', 'nelson', 'rotorua', 'new plymouth',
                        'whangarei', 'invercargill', 'upper hutt', 'lower hutt', 'porirua']
            
            for line in lines:
                line_lower = line.lower()
                for city in nz_cities:
                    if city in line_lower:
                        location_text = line
                        break
                if location_text != 'N/A':
                    break
            
            data['location'] = location_text
            
            # Extract additional car details from title/description with improved patterns
            transmission = 'N/A'
            fuel_type = 'N/A'
            body_style = 'N/A'
            
            # Enhanced transmission detection with more accurate patterns
            transmission_patterns = [
                # Explicit transmission words
                r'\b(manual|automatic|auto|cvt)\b',
                # Gear speed patterns
                r'\b(\d+)\s*speed\s*(manual|auto|automatic)\b',
                # Gear codes (6A, 6M, etc.)
                r'\b(\d+[am])\b',
                # Common abbreviations
                r'\b(6a|6m|5a|5m|4a|4m|6sp|5sp)\b'
            ]
            
            for pattern in transmission_patterns:
                trans_match = re.search(pattern, full_text, re.IGNORECASE)
                if trans_match:
                    trans_text = trans_match.group(1).lower()
                    
                    # Check for explicit transmission types first
                    if 'manual' in trans_text:
                        transmission = 'Manual'
                        break
                    elif 'automatic' in trans_text or 'auto' in trans_text:
                        transmission = 'Automatic'
                        break
                    elif 'cvt' in trans_text:
                        transmission = 'CVT'
                        break
                    
                    # Check for gear codes
                    elif re.search(r'\d+m', trans_text):  # 6M, 5M, etc.
                        transmission = 'Manual'
                        break
                    elif re.search(r'\d+a', trans_text):  # 6A, 5A, etc.
                        transmission = 'Automatic'
                        break
                    elif 'm' in trans_text and not 'auto' in trans_text:
                        transmission = 'Manual'
                        break
                    elif 'a' in trans_text and not 'manual' in trans_text:
                        transmission = 'Automatic'
                        break
            
            # Enhanced fuel type detection with more patterns
            fuel_patterns = [
                # Explicit fuel types
                r'\b(petrol|diesel|hybrid|electric|gasoline|unleaded)\b',
                # Engine displacement codes
                r'\b(2\.0p|2\.0d|2\.0l|2\.0i|1\.8p|1\.6p|2\.4p|2\.5p)\b',
                # Single letter codes in context
                r'\b(2\.0\s*p|2\.0\s*d|2\.0\s*l|2\.0\s*i)\b',
                # Engine type indicators
                r'\b(na|turbo|supercharged)\b'
            ]
            
            for pattern in fuel_patterns:
                fuel_match = re.search(pattern, full_text, re.IGNORECASE)
                if fuel_match:
                    fuel_text = fuel_match.group(1).lower()
                    if 'petrol' in fuel_text or 'gasoline' in fuel_text or 'unleaded' in fuel_text or 'p' in fuel_text:
                        fuel_type = 'Petrol'
                        break
                    elif 'diesel' in fuel_text or 'd' in fuel_text:
                        fuel_type = 'Diesel'
                        break
                    elif 'hybrid' in fuel_text:
                        fuel_type = 'Hybrid'
                        break
                    elif 'electric' in fuel_text:
                        fuel_type = 'Electric'
                        break
            
            # Enhanced body style detection
            body_patterns = [
                r'\b(coupe|coupe|sedan|hatchback|wagon|suv|convertible|roadster)\b',
                r'\b(2\s*dr|4\s*dr|2\s*door|4\s*door)\b',  # Door indicators
                r'\b(gt|limited|ltd|sport|base)\b'  # Trim levels that might indicate body style
            ]
            
            for pattern in body_patterns:
                body_match = re.search(pattern, full_text, re.IGNORECASE)
                if body_match:
                    body_text = body_match.group(1).lower()
                    if 'coupe' in body_text or '2 dr' in body_text or '2 door' in body_text:
                        body_style = 'Coupe'
                    elif 'sedan' in body_text or '4 dr' in body_text or '4 door' in body_text:
                        body_style = 'Sedan'
                    elif 'hatchback' in body_text:
                        body_style = 'Hatchback'
                    elif 'wagon' in body_text:
                        body_style = 'Wagon'
                    elif 'suv' in body_text:
                        body_style = 'SUV'
                    elif 'convertible' in body_text or 'roadster' in body_text:
                        body_style = 'Convertible'
                    break
            
            # For 86/BRZ, default to Coupe if not found (they're typically coupes)
            if body_style == 'N/A' and ('86' in full_text or 'brz' in full_text.lower()):
                body_style = 'Coupe'
            
            # For 86/BRZ, default to Petrol if not found (they're typically petrol)
            if fuel_type == 'N/A' and ('86' in full_text or 'brz' in full_text.lower()):
                fuel_type = 'Petrol'
            
            data['transmission'] = transmission
            data['fuel_type'] = fuel_type
            data['body_style'] = body_style
            
            # Generate unique ID
            data['ID'] = self.generate_unique_id(data['title'], data['location'], data['year'])
            
            # Set car_model before generating search terms
            data['car_model'] = car_model
            
            # Determine if it's an auction
            is_auction = any(word in full_text.lower() for word in ['auction', 'bid', 'reserve', 'ending'])
            data['is_auction'] = is_auction
            data['price_type'] = 'Auction' if is_auction else 'Buy Now'
            
            # Determine seller type
            is_dealer = any(word in full_text.lower() for word in ['dealer', 'motors', 'cars', 'auto', 'ltd', 'limited'])
            data['seller_type'] = 'Dealer' if is_dealer else 'Private'
            data['is_dealer'] = is_dealer
            
            # Extract listing time information
            listing_time = 'N/A'
            listing_date = 'N/A'
            
            # Look for time patterns like "Listed 2 hours ago", "Listed yesterday", "Listed within the last 7 days"
            for line in lines:
                line_lower = line.lower()
                
                # Check for "Listed within the last 7 days" or similar
                if 'listed within the last 7 days' in line_lower:
                    listing_time = 'Within 7 days'
                    listing_date = datetime.now().strftime('%Y-%m-%d')
                    break
                elif 'listed yesterday' in line_lower:
                    yesterday = datetime.now() - timedelta(days=1)
                    listing_time = 'Yesterday'
                    listing_date = yesterday.strftime('%Y-%m-%d')
                    break
                elif 'listed today' in line_lower:
                    listing_time = 'Today'
                    listing_date = datetime.now().strftime('%Y-%m-%d')
                    break
                elif 'listed' in line_lower and ('hour' in line_lower or 'minute' in line_lower):
                    # Extract time like "Listed 2 hours ago"
                    time_match = re.search(r'listed\s+(\d+)\s+(hour|minute)s?\s+ago', line_lower)
                    if time_match:
                        amount = int(time_match.group(1))
                        unit = time_match.group(2)
                        if unit == 'hour':
                            listing_time = f'{amount} hours ago'
                            listing_date = datetime.now().strftime('%Y-%m-%d')
                        elif unit == 'minute':
                            listing_time = f'{amount} minutes ago'
                            listing_date = datetime.now().strftime('%Y-%m-%d')
                    break
            
            # Extract auction/listing end time information
            auction_end_time = 'N/A'
            auction_end_date = 'N/A'
            listing_end_time = 'N/A'
            listing_end_date = 'N/A'
            
            for line in lines:
                line_lower = line.lower()
                
                # Look for auction end times
                if 'ending' in line_lower or 'ends' in line_lower:
                    # Patterns like "Ending in 2 days", "Ends in 5 hours", "Ending today"
                    if 'ending today' in line_lower:
                        auction_end_time = 'Today'
                        auction_end_date = datetime.now().strftime('%Y-%m-%d')
                    elif 'ending tomorrow' in line_lower:
                        tomorrow = datetime.now() + timedelta(days=1)
                        auction_end_time = 'Tomorrow'
                        auction_end_date = tomorrow.strftime('%Y-%m-%d')
                    elif 'ending in' in line_lower:
                        # Extract "Ending in X days/hours"
                        end_match = re.search(r'ending\s+in\s+(\d+)\s+(day|hour)s?', line_lower)
                        if end_match:
                            amount = int(end_match.group(1))
                            unit = end_match.group(2)
                            if unit == 'day':
                                end_date = datetime.now() + timedelta(days=amount)
                                auction_end_time = f'In {amount} days'
                                auction_end_date = end_date.strftime('%Y-%m-%d')
                            elif unit == 'hour':
                                end_date = datetime.now() + timedelta(hours=amount)
                                auction_end_time = f'In {amount} hours'
                                auction_end_date = end_date.strftime('%Y-%m-%d')
                    elif 'ends in' in line_lower:
                        # Extract "Ends in X days/hours"
                        end_match = re.search(r'ends\s+in\s+(\d+)\s+(day|hour)s?', line_lower)
                        if end_match:
                            amount = int(end_match.group(1))
                            unit = end_match.group(2)
                            if unit == 'day':
                                end_date = datetime.now() + timedelta(days=amount)
                                auction_end_time = f'In {amount} days'
                                auction_end_date = end_date.strftime('%Y-%m-%d')
                            elif unit == 'hour':
                                end_date = datetime.now() + timedelta(hours=amount)
                                auction_end_time = f'In {amount} hours'
                                auction_end_date = end_date.strftime('%Y-%m-%d')
                
                # Look for specific date/time patterns
                # Patterns like "Ends 25 Sep 2024", "Ending 25/09/2024", "Ends Sep 25"
                date_patterns = [
                    r'ends?\s+(\d{1,2}\s+\w{3,9}\s+\d{4})',  # Ends 25 Sep 2024
                    r'ends?\s+(\d{1,2}/\d{1,2}/\d{4})',      # Ends 25/09/2024
                    r'ends?\s+(\w{3,9}\s+\d{1,2})',          # Ends Sep 25
                    r'ending\s+(\d{1,2}\s+\w{3,9}\s+\d{4})', # Ending 25 Sep 2024
                    r'ending\s+(\d{1,2}/\d{1,2}/\d{4})',     # Ending 25/09/2024
                    r'ending\s+(\w{3,9}\s+\d{1,2})'           # Ending Sep 25
                ]
                
                for pattern in date_patterns:
                    date_match = re.search(pattern, line_lower)
                    if date_match:
                        date_str = date_match.group(1)
                        try:
                            # Try different date formats
                            for fmt in ['%d %b %Y', '%d/%m/%Y', '%b %d', '%d %B %Y', '%d-%m-%Y']:
                                try:
                                    parsed_date = datetime.strptime(date_str, fmt)
                                    if fmt == '%b %d':  # Sep 25 - assume current year
                                        parsed_date = parsed_date.replace(year=datetime.now().year)
                                    auction_end_time = parsed_date.strftime('%d %b %Y')
                                    auction_end_date = parsed_date.strftime('%Y-%m-%d')
                                    break
                                except ValueError:
                                    continue
                        except:
                            pass
                        break
                
                # Look for time patterns like "Ends at 2:30 PM", "Ending at 14:30"
                time_patterns = [
                    r'ends?\s+at\s+(\d{1,2}:\d{2}\s*(?:am|pm)?)',
                    r'ending\s+at\s+(\d{1,2}:\d{2}\s*(?:am|pm)?)'
                ]
                
                for pattern in time_patterns:
                    time_match = re.search(pattern, line_lower)
                    if time_match:
                        time_str = time_match.group(1)
                        # If we already have a date, add the time
                        if auction_end_date != 'N/A':
                            auction_end_time = f"{auction_end_date} {time_str}"
                        break
            
            # Additional fields
            data['listing_url'] = 'N/A'
            data['listing_id'] = 'N/A'
            data['listing_time'] = listing_time
            data['listing_date'] = listing_date
            data['auction_end_time'] = auction_end_time
            data['auction_end_date'] = auction_end_date
            data['listing_end_time'] = listing_end_time
            data['listing_end_date'] = listing_end_date
            data['scrape_date'] = datetime.now().strftime('%Y-%m-%d')
            data['scrape_time'] = datetime.now().strftime('%H:%M:%S')
            data['last_seen'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            data['is_active'] = True
            
            # Generate search terms and URLs for finding the original listing
            search_data = self.generate_search_terms(data['title'], data['location'], data['year'], data['brand'], data['car_model'])
            data.update(search_data)
            
            # Add notes column for any unclear data
            if notes:
                data['notes'] = '; '.join(notes)
            else:
                data['notes'] = ''
            
            return data
            
        except Exception as e:
            self.logger.error(f"Error extracting listing data: {e}")
            return None

    def scrape_car_listings(self, car_model, url):
        """Scrape listings for a specific car model"""
        self.logger.info(f"Starting scrape for {car_model}")
        
        driver = webdriver.Chrome(options=self.chrome_options)
        all_listings = []
        
        try:
            # Navigate to the page
            driver.get(url)
            self.logger.info(f"Navigated to: {url}")
            
            # Wait for page to load
            time.sleep(5)
            
            # Look for listings
            listings = driver.find_elements(By.CSS_SELECTOR, '.tm-motors-tier-one-search-card__listing-details-container')
            
            if not listings:
                self.logger.warning(f"No listings found for {car_model}")
                return all_listings
            
            self.logger.info(f"Found {len(listings)} listings for {car_model}")
            
            # Extract data from each listing
            for i, listing in enumerate(listings[:20]):  # Limit to first 20 listings
                try:
                    listing_data = self.extract_listing_data(listing, car_model)
                    if listing_data:
                        all_listings.append(listing_data)
                        self.logger.info(f"Extracted listing {i+1}: {listing_data.get('title', 'N/A')[:50]}...")
                except Exception as e:
                    self.logger.error(f"Error processing listing {i+1}: {e}")
                    continue
            
            self.logger.info(f"Successfully extracted {len(all_listings)} listings for {car_model}")
            
        except Exception as e:
            self.logger.error(f"Error scraping {car_model}: {e}")
        
        finally:
            driver.quit()
        
        return all_listings

    def load_existing_dataset(self):
        """Load existing master dataset if it exists"""
        if os.path.exists(self.master_file):
            try:
                df = pd.read_excel(self.master_file)
                self.logger.info(f"Loaded existing dataset with {len(df)} records")
                return df
            except Exception as e:
                self.logger.error(f"Error loading existing dataset: {e}")
                return pd.DataFrame()
        else:
            self.logger.info("No existing dataset found, creating new one")
            return pd.DataFrame()

    def update_dataset(self, new_data):
        """Update the master dataset with new data, preserving existing information"""
        # Load existing data
        existing_df = self.load_existing_dataset()
        
        if existing_df.empty:
            # Create new dataset
            updated_df = pd.DataFrame(new_data)
        else:
            # Convert new data to DataFrame
            new_df = pd.DataFrame(new_data)
            
            # Create a copy of existing data to work with
            updated_df = existing_df.copy()
            
            # Mark all existing listings as potentially inactive first
            updated_df['is_active'] = False
            
            # Process each new listing
            for _, new_row in new_df.iterrows():
                new_id = new_row['ID']
                
                # Check if this listing already exists
                existing_mask = updated_df['ID'] == new_id
                existing_indices = updated_df[existing_mask].index
                
                if len(existing_indices) > 0:
                    # Listing exists - update it while preserving important data
                    existing_idx = existing_indices[0]
                    
                    # Preserve important historical data
                    preserved_price = updated_df.loc[existing_idx, 'price']
                    preserved_kms = updated_df.loc[existing_idx, 'kms']
                    preserved_listing_date = updated_df.loc[existing_idx, 'listing_date']
                    preserved_listing_time = updated_df.loc[existing_idx, 'listing_time']
                    
                    # Update the existing row with new data
                    updated_df.loc[existing_idx] = new_row
                    
                    # Restore preserved data if new data doesn't have it
                    if pd.isna(new_row['price']) or new_row['price'] == 'N/A' or new_row['price'] == '':
                        updated_df.loc[existing_idx, 'price'] = preserved_price
                    if pd.isna(new_row['kms']) or new_row['kms'] == 'N/A' or new_row['kms'] == '':
                        updated_df.loc[existing_idx, 'kms'] = preserved_kms
                    if pd.isna(new_row['listing_date']) or new_row['listing_date'] == 'N/A':
                        updated_df.loc[existing_idx, 'listing_date'] = preserved_listing_date
                    if pd.isna(new_row['listing_time']) or new_row['listing_time'] == 'N/A':
                        updated_df.loc[existing_idx, 'listing_time'] = preserved_listing_time
                    
                    # Mark as active since we found it again
                    updated_df.loc[existing_idx, 'is_active'] = True
                    
                    # Update last_seen timestamp
                    updated_df.loc[existing_idx, 'last_seen'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                else:
                    # New listing - add it
                    updated_df = pd.concat([updated_df, new_row.to_frame().T], ignore_index=True)
            
            # Remove any duplicate IDs (keep the most recent)
            updated_df = updated_df.drop_duplicates(subset=['ID'], keep='first')
        
        return updated_df

    def clean_and_format_data(self, df):
        """Clean and format data for proper Excel number formatting"""
        try:
            # Create a copy to avoid modifying original
            cleaned_df = df.copy()
            
            # Clean year column - convert to numbers if possible
            cleaned_df['year'] = cleaned_df['year'].apply(lambda x: self.clean_number(x, 'year'))
            
            # Clean mileage column - extract numbers from text like "50,000 km"
            cleaned_df['kms'] = cleaned_df['kms'].apply(lambda x: self.clean_mileage(x))
            
            # Clean price column - extract numbers from text like "$25,000"
            cleaned_df['price'] = cleaned_df['price'].apply(lambda x: self.clean_price(x))
            
            self.logger.info("Data cleaned and formatted for Excel")
            return cleaned_df
            
        except Exception as e:
            self.logger.error(f"Error cleaning data: {e}")
            return df

    def clean_number(self, value, column_type):
        """Clean a number value, return number if valid, otherwise return original"""
        if pd.isna(value) or value == 'N/A' or value == '':
            return value
        
        try:
            # Remove any non-digit characters except decimal point
            cleaned = str(value).replace(',', '').replace(' ', '')
            # Extract just the number part
            number_match = re.search(r'(\d+(?:\.\d+)?)', cleaned)
            if number_match:
                number = float(number_match.group(1))
                if column_type == 'year':
                    # For years, return as integer if it's a reasonable year
                    if 1900 <= number <= 2030:
                        return int(number)
                return number
        except:
            pass
        
        return value

    def clean_mileage(self, value):
        """Clean mileage value - extract number from text like '50,000 km' or just numbers"""
        if pd.isna(value) or value == 'N/A' or value == '':
            return ''  # Return blank for empty/invalid values
        
        try:
            # If it's already a number, return it
            if isinstance(value, (int, float)):
                return int(value)
            
            # If it's a string that's already just a number, convert it
            if str(value).replace(',', '').isdigit():
                return int(str(value).replace(',', ''))
            
            # Look for number followed by 'km' (for cases where extraction didn't work)
            match = re.search(r'([\d,]+)\s*km', str(value), re.IGNORECASE)
            if match:
                # Remove commas and convert to number
                number_str = match.group(1).replace(',', '')
                return int(number_str)
        except:
            pass
        
        return ''  # Return blank if not a valid number

    def clean_price(self, value):
        """Clean price value - extract number from text like '$25,000'"""
        if pd.isna(value) or value == 'N/A' or value == '' or value == '':
            return ''  # Return blank for empty/invalid prices
        
        try:
            # Look for $ followed by number
            match = re.search(r'\$([\d,]+)', str(value))
            if match:
                # Remove commas and convert to number
                number_str = match.group(1).replace(',', '')
                price_num = int(number_str)
                # Only return prices >= $1000
                if price_num >= 1000:
                    return price_num
                else:
                    return ''  # Blank for prices < $1000
        except:
            pass
        
        return ''  # Return blank for any other cases

    def create_listing_sort_value(self, row):
        """Create a sortable value for listing time, prioritizing auction end times"""
        # First try to get auction end time for sorting
        auction_end_time = str(row.get('auction_end_time', 'N/A'))
        auction_end_date = str(row.get('auction_end_date', 'N/A'))
        
        # If we have auction end info, use that for sorting
        if auction_end_time != 'N/A' and auction_end_time != '':
            try:
                # Handle auction end time patterns
                if 'today' in auction_end_time.lower():
                    return datetime.now().replace(hour=23, minute=59, second=59, microsecond=0)
                elif 'tomorrow' in auction_end_time.lower():
                    tomorrow = datetime.now() + timedelta(days=1)
                    return tomorrow.replace(hour=23, minute=59, second=59, microsecond=0)
                elif 'in' in auction_end_time.lower():
                    # Extract "In X days/hours"
                    time_match = re.search(r'in\s+(\d+)\s+(day|hour)s?', auction_end_time.lower())
                    if time_match:
                        amount = int(time_match.group(1))
                        unit = time_match.group(2)
                        if unit == 'day':
                            return datetime.now() + timedelta(days=amount)
                        elif unit == 'hour':
                            return datetime.now() + timedelta(hours=amount)
                elif auction_end_date != 'N/A' and auction_end_date != '':
                    return datetime.strptime(auction_end_date, '%Y-%m-%d')
            except:
                pass
        
        # Fall back to listing time
        listing_time = str(row.get('listing_time', 'N/A'))
        listing_date = str(row.get('listing_date', 'N/A'))
        
        try:
            # Convert listing time to a sortable datetime
            if listing_time == 'N/A' or listing_time == '':
                return datetime.min  # Put at the end
            
            # Handle different time formats
            if 'minutes ago' in listing_time.lower():
                minutes = int(re.search(r'(\d+)\s+minutes?\s+ago', listing_time.lower()).group(1))
                return datetime.now() - timedelta(minutes=minutes)
            elif 'hours ago' in listing_time.lower():
                hours = int(re.search(r'(\d+)\s+hours?\s+ago', listing_time.lower()).group(1))
                return datetime.now() - timedelta(hours=hours)
            elif listing_time.lower() == 'today':
                return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            elif listing_time.lower() == 'yesterday':
                yesterday = datetime.now() - timedelta(days=1)
                return yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
            elif 'within 7 days' in listing_time.lower():
                # Put 7 days ago listings after recent ones but before older ones
                week_ago = datetime.now() - timedelta(days=7)
                return week_ago.replace(hour=12, minute=0, second=0, microsecond=0)
            else:
                # Try to parse the listing_date if available
                if listing_date != 'N/A' and listing_date != '':
                    return datetime.strptime(listing_date, '%Y-%m-%d')
                else:
                    return datetime.min
        except:
            return datetime.min

    def apply_conditional_formatting(self, filepath):
        """Apply beautiful conditional formatting to the Excel file"""
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Define beautiful colors
            header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")  # Dark slate gray
            header_font = Font(bold=True, color="FFFFFF")  # White bold text
            
            active_id_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")  # Light green for ID column
            inactive_id_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Light gray for inactive ID column
            inactive_row_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")  # Very light gray for inactive rows
            
            # Light brand-specific colors
            toyota_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")  # Very light cream for Toyota
            subaru_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # Very light blue for Subaru
            
            # Special formatting for optimal choices and steals
            steal_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green for steals
            optimal_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Light orange for optimal choices
            avoid_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light pink for avoid
            budget_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Light purple for budget range
            
            # Style the header row (row 1)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set compact row heights for better visibility
            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 15  # Compact row height
            
            # Set default alignment for all data cells
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    column_letter = cell.column_letter
                    
                    # Set alignment based on column type - all centered for compact view
                    if column_letter in ['A', 'C', 'D', 'E', 'H', 'J', 'T']:  # ID, Year, Kms, Price, Is auction, Is dealer, Is active
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif column_letter in ['B', 'F', 'G', 'I', 'L', 'M', 'U', 'V', 'W', 'Y', 'Z', 'AA', 'AB']:  # Text columns
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif column_letter in ['K', 'N', 'O', 'P', 'Q', 'R', 'S', 'X']:  # Title and URL columns (long text)
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)  # No wrap for compact view
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Find the is_active and brand columns
            header_row = 1
            is_active_col = None
            brand_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=header_row, column=col).value == 'is_active':
                    is_active_col = col
                elif ws.cell(row=header_row, column=col).value == 'brand':
                    brand_col = col
            
            # Find additional columns for categorization
            year_col = None
            price_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == 'year':
                    year_col = col
                elif ws.cell(row=1, column=col).value == 'price':
                    price_col = col
            
            # Apply formatting based on is_active, brand, year, and price columns
            for row in range(2, ws.max_row + 1):
                # Get values for this row
                brand_value = None
                year_value = None
                price_value = None
                
                if brand_col:
                    brand_value = ws.cell(row=row, column=brand_col).value
                if year_col:
                    year_value = ws.cell(row=row, column=year_col).value
                if price_col:
                    price_value = ws.cell(row=row, column=price_col).value
                
                # Determine category based on buying guide criteria
                category = self.categorize_listing(year_value, price_value)
                
                # Determine brand-specific fill
                brand_fill = None
                if brand_value == 'Toyota':
                    brand_fill = toyota_fill
                elif brand_value == 'Subaru':
                    brand_fill = subaru_fill
                
                # Apply formatting based on category and is_active column
                if is_active_col:
                    is_active_cell = ws.cell(row=row, column=is_active_col)
                    
                    if is_active_cell.value == True:
                        # Active listing - apply category-based formatting
                        if category == 'steal':
                            # STEAL: 2015-2016, 18-23k budget, 80-140k km
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = steal_fill
                        elif category == 'optimal':
                            # OPTIMAL: 2015-2016, any price, or 2014 in budget
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = optimal_fill
                        elif category == 'budget':
                            # BUDGET: Within 18-23k range but not optimal year
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = budget_fill
                        elif category == 'avoid':
                            # AVOID: 2012-2013 models
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = avoid_fill
                        else:
                            # Default: light green ID column + brand color for brand column
                            ws.cell(row=row, column=1).fill = active_id_fill  # ID column (column 1)
                            if brand_col and brand_fill:
                                ws.cell(row=row, column=brand_col).fill = brand_fill
                    else:
                        # Inactive listing - light gray ID column and slightly lighter row
                        ws.cell(row=row, column=1).fill = inactive_id_fill  # ID column
                        # Make the entire row slightly lighter
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = inactive_row_fill
                else:
                    # No is_active column, just apply brand colors
                    if brand_col and brand_fill:
                        ws.cell(row=row, column=brand_col).fill = brand_fill
            
            # Apply currency formatting to price column
            price_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == 'price':
                    price_col = col
                    break
            
            # Apply number formatting to appropriate columns
            for col in range(1, ws.max_column + 1):
                column_name = ws.cell(row=1, column=col).value
                
                if column_name == 'price':
                    # Currency formatting for price column
                    for row in range(2, ws.max_row + 1):
                        price_cell = ws.cell(row=row, column=col)
                        if isinstance(price_cell.value, (int, float)) and price_cell.value != '' and price_cell.value != '-':
                            price_cell.number_format = '$#,##0'  # Currency format with $ symbol
                
                elif column_name == 'year':
                    # Number formatting for year column
                    for row in range(2, ws.max_row + 1):
                        year_cell = ws.cell(row=row, column=col)
                        if isinstance(year_cell.value, (int, float)):
                            year_cell.number_format = '0'  # Integer format
                
                elif column_name == 'kms':
                    # Number formatting for mileage column
                    for row in range(2, ws.max_row + 1):
                        kms_cell = ws.cell(row=row, column=col)
                        # Apply number formatting to all kms cells (even if they appear as text)
                        if kms_cell.value is not None and kms_cell.value != '':
                            try:
                                # Convert to number if it's a string that looks like a number
                                if isinstance(kms_cell.value, str) and kms_cell.value.replace(',', '').isdigit():
                                    kms_cell.value = int(kms_cell.value.replace(',', ''))
                                kms_cell.number_format = '#,##0'  # Number format with commas
                            except:
                                # If conversion fails, leave as is
                                pass
            
            # Auto-adjust column widths with better formatting
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                # Calculate max length for this column
                for cell in column:
                    try:
                        if cell.value is not None:
                            # For headers, use the header text length
                            if cell.row == 1:
                                max_length = max(max_length, len(str(cell.value)))
                            else:
                                # For data cells, consider the actual content length
                                cell_length = len(str(cell.value))
                                max_length = max(max_length, cell_length)
                    except:
                        pass
                
                # Set minimum widths for different column types
                min_widths = {
                    'A': 8,   # ID column
                    'B': 10,  # Brand column
                    'C': 6,   # Year column
                    'D': 12,  # Kms column
                    'E': 12,  # Price column
                    'F': 20,  # Location column
                    'G': 15,  # Price type column
                    'H': 10,  # Is auction column
                    'I': 15,  # Seller type column
                    'J': 10,  # Is dealer column
                    'K': 30,  # Title column
                    'L': 15,  # Car model column
                    'M': 20,  # Primary search term column
                    'N': 30,  # Primary TradeMe URL column
                    'O': 30,  # Primary Google URL column
                    'P': 25,  # Search terms column
                    'Q': 30,  # TradeMe search URLs column
                    'R': 30,  # Google search URLs column
                    'S': 30,  # Google Images URLs column
                    'T': 10,  # Is active column
                    'U': 15,  # Last seen column
                    'V': 15,  # Scrape date column
                    'W': 15,  # Scrape time column
                    'X': 50,  # Listing URL column
                    'Y': 15,  # Listing ID column
                    'Z': 12,  # Transmission column
                    'AA': 12, # Fuel type column
                    'AB': 15  # Body style column
                }
                
                # Get the minimum width for this column
                min_width = min_widths.get(column_letter, 10)
                
                # Calculate final width (max of calculated length + 2, minimum width, but cap at 60)
                final_width = max(max_length + 2, min_width)
                final_width = min(final_width, 60)  # Cap at 60 characters
                
                ws.column_dimensions[column_letter].width = final_width
                
                # No text wrapping for compact view - keep all rows same height
                # if column_letter in ['K', 'Q']:  # Title and URL columns
                #     for cell in column:
                #         cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Add legend for color coding
            self.add_color_legend(ws)
            
            wb.save(filepath)
            self.logger.info("Applied beautiful conditional formatting to Excel file")
            
        except Exception as e:
            self.logger.error(f"Error applying conditional formatting: {e}")

    def add_color_legend(self, ws):
        """Add a color legend to explain the formatting"""
        try:
            # Find the last row with data
            last_row = ws.max_row
            
            # Add legend starting 2 rows after the data
            legend_start_row = last_row + 3
            
            # Legend title
            ws.cell(row=legend_start_row, column=1, value="COLOR CODING LEGEND:")
            ws.cell(row=legend_start_row, column=1).font = Font(bold=True, size=12)
            
            # Legend entries
            legend_entries = [
                ("🟢 STEAL", "2015-2016 models in 18-23k budget - PERFECT CHOICE!"),
                ("🟠 OPTIMAL", "2015-2016 models (any price) or 2014 in budget - GREAT CHOICE!"),
                ("🟣 BUDGET", "2014 models in 18-23k budget - BACKUP OPTION"),
                ("🔴 AVOID", "2012-2013 models - Valve spring recall issues!"),
                ("⚪ OTHER", "Other years or no data - Check individually")
            ]
            
            for i, (color, description) in enumerate(legend_entries):
                row = legend_start_row + 1 + i
                ws.cell(row=row, column=1, value=color)
                ws.cell(row=row, column=2, value=description)
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2).font = Font(size=10)
            
            # Add buying guide reference
            guide_row = legend_start_row + len(legend_entries) + 2
            ws.cell(row=guide_row, column=1, value="Based on Toyota 86/Subaru BRZ Reliability-Focused Buying Guide")
            ws.cell(row=guide_row, column=1).font = Font(italic=True, size=9)
            
        except Exception as e:
            self.logger.error(f"Error adding color legend: {e}")

    def save_master_dataset(self, df):
        """Save the master dataset with proper formatting"""
        if df.empty:
            self.logger.warning("No data to save")
            return
        
        # Reorder columns as requested: ID, brand, year, kms, price, location
        column_order = [
            'ID', 'brand', 'year', 'kms', 'price', 'location', 'price_type', 
            'is_auction', 'seller_type', 'is_dealer', 'title', 'car_model', 
            'primary_search_term', 'primary_trademe_url', 'primary_google_url',
            'search_terms', 'trademe_search_urls', 'google_search_urls', 'google_images_urls',
            'listing_time', 'listing_date', 'auction_end_time', 'auction_end_date', 
            'listing_end_time', 'listing_end_date', 'is_active', 'last_seen', 'scrape_date', 'scrape_time', 
            'listing_url', 'listing_id', 'transmission', 'fuel_type', 'body_style', 'notes'
        ]
        
        # Only keep columns that exist in the data
        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]
        
        # Create a sortable listing time column for sorting
        df['listing_sort'] = df.apply(self.create_listing_sort_value, axis=1)
        
        # Create separate sort columns for active and inactive listings
        # Active listings: newest first (descending)
        # Inactive listings: oldest first (ascending) - so oldest inactive at bottom
        df['active_sort'] = df.apply(lambda row: row['listing_sort'] if row['is_active'] == True else datetime.min, axis=1)
        df['inactive_sort'] = df.apply(lambda row: row['listing_sort'] if row['is_active'] == False else datetime.max, axis=1)
        
        # Sort by: 
        # 1. Active listings first (newest first)
        # 2. Inactive listings last (oldest first)
        # 3. Scrape date (most recent first)
        # 4. Last seen (most recent first)
        df = df.sort_values(['is_active', 'active_sort', 'inactive_sort', 'scrape_date', 'last_seen'], 
                           ascending=[False, False, True, False, False])
        
        # Remove the temporary sorting columns
        df = df.drop(['listing_sort', 'active_sort', 'inactive_sort'], axis=1)
        
        # Clean and format data for proper Excel number formatting
        df = self.clean_and_format_data(df)
        
        # Save to Excel (no timestamped copies, just update the master file)
        filepath = self.master_file
        
        try:
            # Create timestamped filename for daily backup
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            daily_backup_file = os.path.join(self.daily_backups_dir, f"86_BRZ_dataset_{timestamp}.xlsx")
            
            # Save main file
            df.to_excel(filepath, index=False, engine='openpyxl')
            self.logger.info(f"86/BRZ dataset saved to: {filepath}")
            
            # Apply conditional formatting to main file
            self.apply_conditional_formatting(filepath)
            
            # Save daily timestamped backup
            df.to_excel(daily_backup_file, index=False, engine='openpyxl')
            self.logger.info(f"Daily backup saved to: {daily_backup_file}")
            
            # Apply conditional formatting to daily backup
            self.apply_conditional_formatting(daily_backup_file)
            
            # Save backup copy to OneDrive
            df.to_excel(self.onedrive_file, index=False, engine='openpyxl')
            self.logger.info(f"86/BRZ dataset backup saved to: {self.onedrive_file}")
            
            # Apply conditional formatting to OneDrive backup
            self.apply_conditional_formatting(self.onedrive_file)
            
            # Print summary
            print(f"\n=== 86/BRZ Dataset Summary ===")
            print(f"Total listings: {len(df)}")
            print(f"Active listings: {len(df[df['is_active'] == True])}")
            print(f"Inactive listings: {len(df[df['is_active'] == False])}")
            for model in df['car_model'].unique():
                count = len(df[df['car_model'] == model])
                print(f"{model}: {count} listings")
            print(f"Main dataset: {self.master_file}")
            print(f"Daily backup: {daily_backup_file}")
            print(f"OneDrive backup: {self.onedrive_file}")
            
        except Exception as e:
            self.logger.error(f"Error saving master dataset: {e}")

    def scrape_all_cars(self):
        """Scrape listings for all car models"""
        all_data = []
        
        for car_model, url in self.urls.items():
            listings = self.scrape_car_listings(car_model, url)
            all_data.extend(listings)
            time.sleep(2)  # Delay between requests
        
        return all_data

    def run(self):
        """Main execution method"""
        self.logger.info("Starting 86/BRZ Dataset Scraper")
        start_time = datetime.now()
        
        try:
            # Scrape all car data
            new_data = self.scrape_all_cars()
            
            # Update master dataset
            updated_df = self.update_dataset(new_data)
            
            # Save master dataset
            self.save_master_dataset(updated_df)
            
            end_time = datetime.now()
            duration = end_time - start_time
            self.logger.info(f"Master dataset update completed in {duration}")
            
        except Exception as e:
            self.logger.error(f"Error during master dataset update: {e}")

def main():
    """Main function to run the master scraper"""
    scraper = StreamlinedMasterScraper()
    scraper.run()

if __name__ == "__main__":
    main()
